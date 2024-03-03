#!/usr/bin/env python
# coding: utf-8

# In[1]:


from bs4 import BeautifulSoup
import requests


# In[2]:


from openpyxl import Workbook


# In[ ]:


excel = Workbook()


# In[ ]:


print(excel.sheetnames)


# In[ ]:


# want sheet as active because there may be 3 sheets in excel
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)


# In[ ]:


sheet.append(['Rank','Movie Name','Year Of Release','Release Date','Genre','Runtime (in minutes)','Rating (5)', 'Country(s)','IMDB Rating (10)'])

#movie_no, movie_title, movie_year, release_date, movie_genre, movie_rt, movie_rating, movie_country, imdb_rating


# In[2]:


try:
    source = requests.get('https://www.moviemeter.com/movies/top-250-best-movies-of-all-time')
    source.raise_for_status()
    
    soup = BeautifulSoup(source.text,'html.parser')
    #print(soup)
    tbody_movies = soup.find_all('tbody')
    
    all_tr_movies = []
    all_links =[]
    
    count = 0
    
    for tbody_movie in tbody_movies:
        tr_movie = tbody_movie.find_all('tr')
        
        all_tr_movies.extend(tr_movie)
        
        
        
    for all_tr_movie in all_tr_movies:
        
        count += 1
        href = all_tr_movie.find('td').a.get('href')
        
        movie_no = all_tr_movie.find('td').span.text
        
        movie_genre = all_tr_movie.find_all('div', class_='sub')[-2].text
        
        movie_rt = all_tr_movie.find_all('div', class_='sub')[-1].text.split(' ')[0]
        
        linked_page_url = 'https://www.moviemeter.com' + href
        # Send a new HTTP request to the linked page
        linked_page_source = requests.get(linked_page_url)
        linked_page_source.raise_for_status()
        # Parse the HTML content of the linked page
        linked_page_soup = BeautifulSoup(linked_page_source.text, 'html.parser')
        
        movie_title = linked_page_soup.find('div', class_='figure').a.get('title')
        
        movie_year = linked_page_soup.find('div', class_='title movie').h1.text[-5:].strip(')')
        
        movie_rating = all_tr_movie.find('div', class_=lambda x: x and x.startswith('mm_star')).get_text(strip=True)[:4].replace(',', '.')
        ##movie rating from linked page is below
        #movie_rating = linked_page_soup.find('div', class_='rating').span.get_text(strip=True)[:4].replace(',', '.')

        movie_country = linked_page_soup.find_all('h3', class_ = 'h3-film--info oorsprong')[-1].span.get_text(strip=True)
        
        imdb_rating = linked_page_soup.find('div', class_='sub_3').find_all('span', class_ = 'h3-nobold')[-2].a.text[:3].replace(',', '.')
        
        release_date = linked_page_soup.find('div',class_='sub_3').find_all('span', class_ = 'h3-nobold')[-1].text

        print(movie_no, movie_title, movie_year, release_date, movie_genre, movie_rt, movie_rating, movie_country, imdb_rating)
        #sheet.append([movie_no, movie_title, movie_year, release_date, movie_genre, movie_rt, movie_rating, movie_country, imdb_rating])
        #break
        if count==4:
            break
        
except Exception as e:
    print(e,'not worked')


# In[ ]:


excel.save('Top 25 Best Movies of All Time from Movie Meter(updated).xlsx')


# In[3]:


import os
#tells the directory
print(os.getcwd())


# In[ ]:





# making better var after some mistakes
# '''try:
#     source = requests.get('https://www.moviemeter.com/movies/drama/the-shawshank-redemption')
#     source.raise_for_status()
#     
#     soup = BeautifulSoup(source.text,'html.parser')
#     
#     release_date = soup.find('div',class_='sub_3').find_all('span', class_ = 'h3-nobold')[-1].text
#     print(len(release_date))
#     print(release_date)    
#         
# except Exception as e:
#     print(e,'not worked')'''

# making better var after some mistakes
# try:
#     source = requests.get('https://www.moviemeter.com/movies/drama/der-kommer-en-dag')
#     source.raise_for_status()
#     
#     soup = BeautifulSoup(source.text,'html.parser')
#     
#     imdb_rating = soup.find('div', class_='sub_3').find_all('span', class_ = 'h3-nobold')[-2].a.text[:3].replace(',', '.')
#     release_date = soup.find('div',class_='sub_3').find_all('span', class_ = 'h3-nobold')[-1].text
#     print(len(imdb_rating))
#     print(imdb_rating)
# except Exception as e:
#     print(e,'not worked')

# In[ ]:





# In[4]:


tbody_movies


# In[ ]:




